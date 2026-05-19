from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("asian_games_lol_selection_data.xlsx")
OUT = Path("weight_scenarios_5_40.csv")

COMPONENTS = [
    ("league", "league_score_avg"),
    ("worlds", "worlds_score_avg"),
    ("kespa", "kespa_score_avg"),
    ("allpro", "allpro_norm_0_100"),
    ("pog_pom", "pog_pom_norm_0_100"),
]

GROUPS = {
    "Top": ["Zeus", "Kiin", "Doran"],
    "Jungle": ["Canyon", "Oner"],
    "Bot": ["Gumayusi", "Peyz", "Viper", "Ruler"],
}


def read_summary() -> dict[str, dict]:
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook["Candidate_Summary"]
    headers = [cell.value for cell in sheet[1]]
    return {
        row[headers.index("player")]: dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
    }


def build_vectors(rows: dict[str, dict]) -> dict[str, list[float]]:
    players = sorted({player for group in GROUPS.values() for player in group})
    return {
        player: [rows[player][field] for _, field in COMPONENTS]
        for player in players
    }


def score(player: str, weights: tuple[int, ...], vectors: dict[str, list[float]]) -> float:
    return sum(value * weight / 100 for value, weight in zip(vectors[player], weights))


def ranked(
    players: list[str],
    weights: tuple[int, ...],
    vectors: dict[str, list[float]],
) -> list[tuple[str, float]]:
    return sorted(
        ((player, score(player, weights, vectors)) for player in players),
        key=lambda item: (-item[1], item[0]),
    )


def main() -> None:
    rows = read_summary()
    vectors = build_vectors(rows)
    fieldnames = [
        "league",
        "worlds",
        "kespa",
        "allpro",
        "pog_pom",
        "top_winner",
        "top_gap",
        "zeus_score",
        "kiin_score",
        "doran_score",
        "jungle_winner",
        "jungle_gap",
        "canyon_score",
        "oner_score",
        "bot_winner",
        "bot_gap",
        "gumayusi_score",
        "peyz_score",
        "viper_score",
        "ruler_score",
        "selected_zeus_canyon_guma",
    ]

    total = 0
    selected_total = 0
    position_counts = {group: Counter() for group in GROUPS}
    combo_counts: Counter[tuple[str, str, str]] = Counter()

    with OUT.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        for league in range(5, 41):
            for worlds in range(5, 41):
                for kespa in range(5, 41):
                    for allpro in range(5, 41):
                        pog_pom = 100 - league - worlds - kespa - allpro
                        if pog_pom < 5 or pog_pom > 40:
                            continue

                        weights = (league, worlds, kespa, allpro, pog_pom)
                        top = ranked(GROUPS["Top"], weights, vectors)
                        jungle = ranked(GROUPS["Jungle"], weights, vectors)
                        bot = ranked(GROUPS["Bot"], weights, vectors)
                        winners = (top[0][0], jungle[0][0], bot[0][0])
                        selected = winners == ("Zeus", "Canyon", "Gumayusi")

                        total += 1
                        selected_total += int(selected)
                        for group_name, group_ranked in [
                            ("Top", top),
                            ("Jungle", jungle),
                            ("Bot", bot),
                        ]:
                            position_counts[group_name][group_ranked[0][0]] += 1
                        combo_counts[winners] += 1

                        writer.writerow(
                            {
                                "league": league,
                                "worlds": worlds,
                                "kespa": kespa,
                                "allpro": allpro,
                                "pog_pom": pog_pom,
                                "top_winner": top[0][0],
                                "top_gap": round(top[0][1] - top[1][1], 4),
                                "zeus_score": round(score("Zeus", weights, vectors), 4),
                                "kiin_score": round(score("Kiin", weights, vectors), 4),
                                "doran_score": round(score("Doran", weights, vectors), 4),
                                "jungle_winner": jungle[0][0],
                                "jungle_gap": round(jungle[0][1] - jungle[1][1], 4),
                                "canyon_score": round(score("Canyon", weights, vectors), 4),
                                "oner_score": round(score("Oner", weights, vectors), 4),
                                "bot_winner": bot[0][0],
                                "bot_gap": round(bot[0][1] - bot[1][1], 4),
                                "gumayusi_score": round(score("Gumayusi", weights, vectors), 4),
                                "peyz_score": round(score("Peyz", weights, vectors), 4),
                                "viper_score": round(score("Viper", weights, vectors), 4),
                                "ruler_score": round(score("Ruler", weights, vectors), 4),
                                "selected_zeus_canyon_guma": "yes" if selected else "no",
                            }
                        )

    print(f"wrote {OUT.resolve()}")
    print(f"total scenarios: {total}")
    print(f"Zeus/Canyon/Gumayusi scenarios: {selected_total}")
    print("winner counts:")
    for group_name, counts in position_counts.items():
        print(f"- {group_name}: {dict(counts.most_common())}")
    print("winner combos:")
    for winners, count in combo_counts.most_common():
        print(f"- {winners}: {count}")


if __name__ == "__main__":
    main()

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("asian_games_lol_selection_data.xlsx")
OUT = Path("selection_dashboard.html")
PAGES_OUT = Path("index.html")

DEFAULT_WEIGHTS = {
    "league": 17,
    "worlds": 26,
    "kespa": 5,
    "allpro": 5,
    "pog_pom": 47,
}

COMPONENTS = [
    {"key": "league", "label": "리그", "field": "league_score_avg"},
    {"key": "worlds", "label": "월즈", "field": "worlds_score_avg"},
    {"key": "kespa", "label": "KeSPA", "field": "kespa_score_avg"},
    {"key": "allpro", "label": "All-Pro", "field": "allpro_norm_0_100"},
    {"key": "pog_pom", "label": "POG/POM", "field": "pog_pom_norm_0_100"},
]

GROUPS = [
    {"key": "top", "label": "Top", "players": ["Zeus", "Kiin"]},
    {"key": "jungle", "label": "Jungle", "players": ["Canyon", "Oner"]},
    {"key": "bot", "label": "Bot", "players": ["Gumayusi", "Peyz", "Viper"]},
]

PLAYER_COLORS = {
    "Zeus": "#d1495b",
    "Kiin": "#2a9d8f",
    "Canyon": "#264653",
    "Oner": "#e9a13b",
    "Gumayusi": "#e76f51",
    "Peyz": "#457b9d",
    "Viper": "#6a994e",
}

REQUIRED_PLAYERS = {player for group in GROUPS for player in group["players"]}
VALIDATION_AREAS = {"LPL regular rows", "LPL POG/POM", "POG/POM scope", "Detailed stats"}


def cell_float(value: object) -> float:
    if value is None:
        return 0.0
    return float(value)


def read_workbook() -> dict:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {WORKBOOK}")

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        summary = workbook["Candidate_Summary"]
        rows = list(summary.iter_rows(values_only=True))
        headers = {name: index for index, name in enumerate(rows[0])}

        players = {}
        for row in rows[1:]:
            player = row[headers["player"]]
            if player not in REQUIRED_PLAYERS:
                continue
            scores = {}
            for component in COMPONENTS:
                scores[component["key"]] = round(cell_float(row[headers[component["field"]]]), 2)
            players[player] = {
                "name": player,
                "scores": scores,
                "color": PLAYER_COLORS[player],
            }

        missing = sorted(REQUIRED_PLAYERS - set(players))
        if missing:
            raise ValueError(f"Missing required players from Candidate_Summary: {', '.join(missing)}")

        validation = []
        report = workbook["Validation_Report"]
        report_rows = list(report.iter_rows(values_only=True))
        report_headers = {name: index for index, name in enumerate(report_rows[0])}
        for row in report_rows[1:]:
            area = row[report_headers["area"]]
            if area in VALIDATION_AREAS:
                validation.append(
                    {
                        "area": area,
                        "status": row[report_headers["status"]],
                        "finding": row[report_headers["finding"]],
                        "action": row[report_headers["action"]],
                    }
                )

        return {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "components": COMPONENTS,
            "defaultWeights": DEFAULT_WEIGHTS,
            "groups": GROUPS,
            "players": players,
            "validation": validation,
            "botWarning": "원딜 비교는 LPL POG/POM/상세지표 미완성으로 참고용입니다.",
        }
    finally:
        workbook.close()


def weighted_score(player: dict, weights: dict[str, int]) -> float:
    total = 0.0
    for component in COMPONENTS:
        key = component["key"]
        total += player["scores"][key] * weights[key] / 100
    return total


def default_results(data: dict) -> list[str]:
    lines = []
    for group in GROUPS:
        scored = [
            (player, weighted_score(data["players"][player], DEFAULT_WEIGHTS))
            for player in group["players"]
        ]
        scored.sort(key=lambda item: item[1], reverse=True)
        winner, winner_score = scored[0]
        runner_score = scored[1][1] if len(scored) > 1 else winner_score
        lines.append(
            f"{group['label']}: {winner} {winner_score:.2f} (gap {winner_score - runner_score:.2f})"
        )
    return lines


HTML_TEMPLATE = """<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>선발 가중치 대시보드</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f5f6f4;
      --panel: #ffffff;
      --text: #202326;
      --muted: #626970;
      --line: #d9dedb;
      --soft: #eef2ef;
      --accent: #0f8b8d;
      --danger: #c44536;
      --shadow: 0 12px 28px rgba(32, 35, 38, 0.08);
    }

    * {
      box-sizing: border-box;
    }

    body {
      margin: 0;
      min-width: 320px;
      background: var(--bg);
      color: var(--text);
      font-family: Arial, "Apple SD Gothic Neo", "Noto Sans KR", sans-serif;
      letter-spacing: 0;
    }

    button,
    input {
      font: inherit;
    }

    .shell {
      width: min(1180px, calc(100% - 32px));
      margin: 0 auto;
      padding: 28px 0 36px;
    }

    header {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 18px;
    }

    h1 {
      margin: 0;
      font-size: 28px;
      line-height: 1.15;
      font-weight: 800;
    }

    .meta {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
      text-align: right;
    }

    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }

    .controls-panel {
      padding: 18px;
      margin-bottom: 16px;
    }

    .toolbar {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 14px;
      margin-bottom: 16px;
    }

    .weight-total {
      min-width: 112px;
      color: var(--muted);
      font-size: 14px;
      font-weight: 700;
    }

    .weight-total strong {
      color: var(--accent);
      font-size: 18px;
    }

    .buttons {
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      justify-content: flex-end;
    }

    .btn {
      border: 1px solid var(--line);
      background: #ffffff;
      color: var(--text);
      border-radius: 8px;
      padding: 9px 12px;
      cursor: pointer;
      font-weight: 700;
    }

    .btn:hover,
    .btn:focus-visible {
      border-color: var(--accent);
      outline: none;
    }

    .controls {
      display: grid;
      grid-template-columns: repeat(5, minmax(130px, 1fr));
      gap: 14px;
    }

    .control {
      min-width: 0;
      padding: 12px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--soft);
    }

    .control-head {
      display: flex;
      justify-content: space-between;
      gap: 10px;
      margin-bottom: 12px;
      font-weight: 800;
      font-size: 14px;
    }

    .control-value {
      color: var(--accent);
      white-space: nowrap;
    }

    input[type="range"] {
      width: 100%;
      accent-color: var(--accent);
    }

    .summary-grid {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 14px;
      margin-bottom: 16px;
    }

    .summary-card {
      padding: 14px;
      min-height: 92px;
    }

    .summary-label {
      color: var(--muted);
      font-size: 13px;
      font-weight: 800;
      text-transform: uppercase;
    }

    .summary-winner {
      margin-top: 8px;
      font-size: 22px;
      font-weight: 900;
      overflow-wrap: anywhere;
    }

    .summary-detail {
      margin-top: 4px;
      color: var(--muted);
      font-size: 13px;
    }

    .groups {
      display: grid;
      gap: 16px;
    }

    .group-card {
      padding: 18px;
    }

    .group-head {
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 16px;
    }

    .group-title {
      margin: 0;
      font-size: 20px;
      line-height: 1.2;
    }

    .group-gap {
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
      text-align: right;
    }

    .candidate-list {
      display: grid;
      gap: 12px;
    }

    .candidate-row {
      display: grid;
      grid-template-columns: minmax(118px, 160px) 1fr minmax(74px, auto);
      align-items: center;
      gap: 12px;
      padding: 10px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: #ffffff;
    }

    .candidate-row.is-winner {
      border-color: color-mix(in srgb, var(--row-color), #ffffff 35%);
      background: color-mix(in srgb, var(--row-color), #ffffff 92%);
    }

    .candidate-name {
      min-width: 0;
      font-size: 16px;
      font-weight: 900;
      overflow-wrap: anywhere;
    }

    .candidate-rank {
      display: inline-flex;
      justify-content: center;
      align-items: center;
      width: 26px;
      height: 26px;
      margin-right: 8px;
      border-radius: 50%;
      background: var(--soft);
      color: var(--muted);
      font-size: 12px;
      font-weight: 900;
      vertical-align: middle;
    }

    .is-winner .candidate-rank {
      background: var(--row-color);
      color: #ffffff;
    }

    .bar-track {
      position: relative;
      width: 100%;
      height: 18px;
      overflow: hidden;
      border-radius: 8px;
      background: #e7ebe8;
    }

    .bar-fill {
      width: 0;
      height: 100%;
      border-radius: 8px;
      background: var(--row-color);
      transition: width 160ms ease;
    }

    .candidate-score {
      text-align: right;
      font-size: 18px;
      font-weight: 900;
      font-variant-numeric: tabular-nums;
    }

    .components-table {
      margin-top: 14px;
      overflow-x: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }

    table {
      width: 100%;
      min-width: 560px;
      border-collapse: collapse;
      font-size: 13px;
    }

    th,
    td {
      padding: 9px 10px;
      border-bottom: 1px solid var(--line);
      text-align: right;
      font-variant-numeric: tabular-nums;
    }

    th:first-child,
    td:first-child {
      text-align: left;
      font-weight: 800;
    }

    tr:last-child td {
      border-bottom: 0;
    }

    th {
      background: var(--soft);
      color: var(--muted);
      font-weight: 800;
    }

    .warning {
      margin-top: 16px;
      padding: 14px 16px;
      border: 1px solid color-mix(in srgb, var(--danger), #ffffff 55%);
      border-radius: 8px;
      background: color-mix(in srgb, var(--danger), #ffffff 92%);
    }

    .warning-title {
      margin: 0 0 8px;
      color: var(--danger);
      font-weight: 900;
    }

    .warning ul {
      margin: 0;
      padding-left: 18px;
      color: #4a3430;
      font-size: 13px;
      line-height: 1.55;
    }

    @media (max-width: 900px) {
      header,
      .toolbar,
      .group-head {
        align-items: stretch;
        flex-direction: column;
      }

      .meta,
      .group-gap {
        text-align: left;
      }

      .controls,
      .summary-grid {
        grid-template-columns: 1fr;
      }
    }

    @media (max-width: 640px) {
      .shell {
        width: min(100% - 20px, 1180px);
        padding-top: 18px;
      }

      h1 {
        font-size: 24px;
      }

      .candidate-row {
        grid-template-columns: 1fr;
      }

      .candidate-score {
        text-align: left;
      }
    }
  </style>
</head>
<body>
  <script type="application/json" id="dashboard-data">__DASHBOARD_DATA__</script>
  <main class="shell">
    <header>
      <div>
        <h1>선발 가중치 대시보드</h1>
      </div>
      <div class="meta">
        <div>데이터: asian_games_lol_selection_data.xlsx</div>
        <div id="generated-at"></div>
      </div>
    </header>

    <section class="panel controls-panel" aria-label="가중치 조절">
      <div class="toolbar">
        <div class="weight-total">합계 <strong id="weight-total">100%</strong></div>
        <div class="buttons">
          <button class="btn" id="reset-default" type="button">기본값으로 복원</button>
          <button class="btn" id="reset-even" type="button">균등</button>
        </div>
      </div>
      <div class="controls" id="controls"></div>
    </section>

    <section class="summary-grid" id="summary-grid" aria-label="포지션별 승자 요약"></section>
    <section class="groups" id="groups" aria-label="포지션별 비교 그래프"></section>

    <section class="warning" aria-label="데이터 주의">
      <p class="warning-title" id="warning-title"></p>
      <ul id="warnings"></ul>
    </section>
  </main>

  <script>
    const DATA = JSON.parse(document.getElementById("dashboard-data").textContent);
    const componentKeys = DATA.components.map((component) => component.key);
    let weights = { ...DATA.defaultWeights };

    const formatScore = new Intl.NumberFormat("ko-KR", {
      minimumFractionDigits: 2,
      maximumFractionDigits: 2,
    });

    const controlsEl = document.getElementById("controls");
    const groupsEl = document.getElementById("groups");
    const summaryEl = document.getElementById("summary-grid");
    const totalEl = document.getElementById("weight-total");

    function clampWeight(value) {
      const parsed = Number(value);
      if (!Number.isFinite(parsed)) return 0;
      return Math.max(0, Math.min(100, Math.round(parsed)));
    }

    function distributeRemainder(changedKey, changedValue) {
      const next = { ...weights, [changedKey]: clampWeight(changedValue) };
      const remaining = 100 - next[changedKey];
      const rest = componentKeys.filter((key) => key !== changedKey);
      const oldTotal = rest.reduce((sum, key) => sum + weights[key], 0);
      const exact = rest.map((key) => (oldTotal > 0 ? (weights[key] / oldTotal) * remaining : remaining / rest.length));
      const floors = exact.map((value) => Math.floor(value));
      let leftover = remaining - floors.reduce((sum, value) => sum + value, 0);
      const order = exact
        .map((value, index) => ({ index, fraction: value - Math.floor(value) }))
        .sort((a, b) => b.fraction - a.fraction);

      let pointer = 0;
      while (leftover > 0) {
        floors[order[pointer % order.length].index] += 1;
        leftover -= 1;
        pointer += 1;
      }

      rest.forEach((key, index) => {
        next[key] = floors[index];
      });
      weights = next;
    }

    function setWeights(nextWeights) {
      weights = { ...nextWeights };
      update();
    }

    function playerScore(playerName) {
      const scores = DATA.players[playerName].scores;
      return componentKeys.reduce((sum, key) => sum + scores[key] * weights[key] / 100, 0);
    }

    function groupResult(group) {
      const ranked = group.players
        .map((playerName) => ({
          name: playerName,
          score: playerScore(playerName),
          color: DATA.players[playerName].color,
          scores: DATA.players[playerName].scores,
        }))
        .sort((a, b) => b.score - a.score);
      const gap = ranked.length > 1 ? ranked[0].score - ranked[1].score : 0;
      return { ranked, winner: ranked[0], gap };
    }

    function renderControls() {
      controlsEl.replaceChildren();
      DATA.components.forEach((component) => {
        const wrap = document.createElement("label");
        wrap.className = "control";
        wrap.htmlFor = `weight-${component.key}`;

        const head = document.createElement("div");
        head.className = "control-head";

        const label = document.createElement("span");
        label.textContent = component.label;

        const value = document.createElement("span");
        value.className = "control-value";
        value.id = `value-${component.key}`;

        const input = document.createElement("input");
        input.id = `weight-${component.key}`;
        input.type = "range";
        input.min = "0";
        input.max = "100";
        input.step = "1";
        input.setAttribute("aria-label", `${component.label} 가중치`);
        input.addEventListener("input", (event) => {
          distributeRemainder(component.key, event.target.value);
          update();
        });

        head.append(label, value);
        wrap.append(head, input);
        controlsEl.append(wrap);
      });
    }

    function renderStaticGroups() {
      groupsEl.replaceChildren();
      DATA.groups.forEach((group) => {
        const section = document.createElement("section");
        section.className = "panel group-card";
        section.id = `group-${group.key}`;

        const head = document.createElement("div");
        head.className = "group-head";

        const title = document.createElement("h2");
        title.className = "group-title";
        title.textContent = group.label;

        const gap = document.createElement("div");
        gap.className = "group-gap";
        gap.id = `gap-${group.key}`;

        const list = document.createElement("div");
        list.className = "candidate-list";
        list.id = `list-${group.key}`;

        const tableWrap = document.createElement("div");
        tableWrap.className = "components-table";

        const table = document.createElement("table");
        table.id = `table-${group.key}`;

        head.append(title, gap);
        tableWrap.append(table);
        section.append(head, list, tableWrap);
        groupsEl.append(section);
      });
    }

    function renderSummary(resultsByGroup) {
      summaryEl.replaceChildren();
      DATA.groups.forEach((group) => {
        const result = resultsByGroup[group.key];
        const card = document.createElement("article");
        card.className = "panel summary-card";

        const label = document.createElement("div");
        label.className = "summary-label";
        label.textContent = group.label;

        const winner = document.createElement("div");
        winner.className = "summary-winner";
        winner.style.color = result.winner.color;
        winner.textContent = result.winner.name;

        const detail = document.createElement("div");
        detail.className = "summary-detail";
        detail.textContent = `${formatScore.format(result.winner.score)}점 · 차이 ${formatScore.format(result.gap)}점`;

        card.append(label, winner, detail);
        summaryEl.append(card);
      });
    }

    function renderCandidateList(group, result) {
      const list = document.getElementById(`list-${group.key}`);
      list.replaceChildren();

      result.ranked.forEach((candidate, index) => {
        const row = document.createElement("div");
        row.className = `candidate-row${index === 0 ? " is-winner" : ""}`;
        row.style.setProperty("--row-color", candidate.color);

        const name = document.createElement("div");
        name.className = "candidate-name";

        const rank = document.createElement("span");
        rank.className = "candidate-rank";
        rank.textContent = String(index + 1);

        const nameText = document.createElement("span");
        nameText.textContent = candidate.name;
        name.append(rank, nameText);

        const track = document.createElement("div");
        track.className = "bar-track";
        const fill = document.createElement("div");
        fill.className = "bar-fill";
        fill.style.width = `${Math.max(0, Math.min(100, candidate.score))}%`;
        track.append(fill);

        const score = document.createElement("div");
        score.className = "candidate-score";
        score.textContent = formatScore.format(candidate.score);

        row.append(name, track, score);
        list.append(row);
      });
    }

    function renderComponentTable(group, result) {
      const table = document.getElementById(`table-${group.key}`);
      table.replaceChildren();

      const thead = document.createElement("thead");
      const headerRow = document.createElement("tr");
      ["선수", ...DATA.components.map((component) => component.label), "총점"].forEach((label) => {
        const th = document.createElement("th");
        th.textContent = label;
        headerRow.append(th);
      });
      thead.append(headerRow);

      const tbody = document.createElement("tbody");
      result.ranked.forEach((candidate) => {
        const tr = document.createElement("tr");
        const playerCell = document.createElement("td");
        playerCell.textContent = candidate.name;
        tr.append(playerCell);

        DATA.components.forEach((component) => {
          const td = document.createElement("td");
          const raw = candidate.scores[component.key];
          const contribution = raw * weights[component.key] / 100;
          td.textContent = `${formatScore.format(raw)} / ${formatScore.format(contribution)}`;
          tr.append(td);
        });

        const total = document.createElement("td");
        total.textContent = formatScore.format(candidate.score);
        tr.append(total);
        tbody.append(tr);
      });

      table.append(thead, tbody);
    }

    function renderWarnings() {
      document.getElementById("warning-title").textContent = DATA.botWarning;
      const list = document.getElementById("warnings");
      list.replaceChildren();
      DATA.validation.forEach((item) => {
        const li = document.createElement("li");
        li.textContent = `${item.area}: ${item.finding}`;
        list.append(li);
      });
    }

    function updateControls() {
      DATA.components.forEach((component) => {
        const input = document.getElementById(`weight-${component.key}`);
        const value = document.getElementById(`value-${component.key}`);
        input.value = String(weights[component.key]);
        value.textContent = `${weights[component.key]}%`;
      });
      const total = componentKeys.reduce((sum, key) => sum + weights[key], 0);
      totalEl.textContent = `${total}%`;
    }

    function update() {
      updateControls();
      const resultsByGroup = {};
      DATA.groups.forEach((group) => {
        const result = groupResult(group);
        resultsByGroup[group.key] = result;
        document.getElementById(`gap-${group.key}`).textContent =
          `${result.winner.name} 우세 · ${formatScore.format(result.gap)}점 차`;
        renderCandidateList(group, result);
        renderComponentTable(group, result);
      });
      renderSummary(resultsByGroup);
    }

    document.getElementById("reset-default").addEventListener("click", () => {
      setWeights(DATA.defaultWeights);
    });

    document.getElementById("reset-even").addEventListener("click", () => {
      setWeights(Object.fromEntries(componentKeys.map((key) => [key, 20])));
    });

    document.getElementById("generated-at").textContent = `생성: ${DATA.generatedAt}`;
    renderControls();
    renderStaticGroups();
    renderWarnings();
    update();
  </script>
</body>
</html>
"""


def render_html(data: dict) -> str:
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return HTML_TEMPLATE.replace("__DASHBOARD_DATA__", payload)


def main() -> None:
    data = read_workbook()
    html = render_html(data)
    OUT.write_text(html, encoding="utf-8")
    PAGES_OUT.write_text(html, encoding="utf-8")
    print(f"wrote {OUT.resolve()}")
    print(f"wrote {PAGES_OUT.resolve()}")
    print("default results:")
    for line in default_results(data):
        print(f"- {line}")


if __name__ == "__main__":
    main()

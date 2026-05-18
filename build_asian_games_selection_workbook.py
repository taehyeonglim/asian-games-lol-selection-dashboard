from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path("asian_games_lol_selection_data.xlsx")


SOURCES = [
    {
        "id": "S1",
        "name": "Inven national team article",
        "url": "https://m.inven.co.kr/webzine/wznews.php?idx=316508",
        "used_for": "Selected roster, selection criteria, selected-player tournament summaries",
        "note": "Reports criteria as regional regular league, KeSPA Cup, Worlds from 2024 onward; MSI/FST/EWC excluded.",
    },
    {
        "id": "S2",
        "name": "LCK 2024 Spring - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LCK/2024_Season/Spring_Season",
        "used_for": "Regular standings, All-Pro, POG standings",
        "note": "",
    },
    {
        "id": "S3",
        "name": "LCK 2024 Summer - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LCK/2024_Season/Summer_Season",
        "used_for": "Regular standings, All-Pro, POG standings",
        "note": "",
    },
    {
        "id": "S4",
        "name": "LCK 2025 Rounds 1-2 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LCK/2025_Season/Rounds_1-2",
        "used_for": "Regular standings and POM standings",
        "note": "POM unit is Times, not 100-point POG.",
    },
    {
        "id": "S5",
        "name": "LCK 2025 Rounds 3-5 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LCK/2025_Season/Rounds_3-5",
        "used_for": "Regular standings, All-Pro, POM standings",
        "note": "",
    },
    {
        "id": "S6",
        "name": "LCK 2026 Rounds 1-2 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LCK/2026_Season/Rounds_1-2",
        "used_for": "Current 2026 LCK standings and POM standings through Week 7",
        "note": "Snapshot used: HLE 12-2, KT 11-3, T1 10-4, Gen.G 10-4.",
    },
    {
        "id": "S7",
        "name": "KeSPA Cup 2025 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/2025_LoL_KeSPA_Cup",
        "used_for": "KeSPA placement, POG standings",
        "note": "",
    },
    {
        "id": "S8",
        "name": "KeSPA Cup 2024 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/2024_LoL_KeSPA_Cup",
        "used_for": "KeSPA placement and participation",
        "note": "",
    },
    {
        "id": "S9",
        "name": "LPL 2025 Split 1 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LPL/2025_Season/Split_1",
        "used_for": "Peyz/JDG 2025 LPL Split 1 regular and All-Pro context",
        "note": "LPL 2025 coverage remains partial until Split 2/3 regular POG and detailed stats are added.",
    },
    {
        "id": "S10",
        "name": "Viper 2026 LPL pages - Leaguepedia/Liquipedia",
        "url": "https://lol.fandom.com/wiki/LPL/2026_Season/Split_1",
        "used_for": "Viper 2026 LPL context",
        "note": "LPL 2026 coverage is partial in this workbook.",
    },
    {
        "id": "S11",
        "name": "Worlds 2024/2025 event pages",
        "url": "https://liquipedia.net/leagueoflegends/World_Championship",
        "used_for": "World Championship placements",
        "note": "",
    },
    {
        "id": "S12",
        "name": "gol.gg LCK 2024 Spring ranking",
        "url": "https://gol.gg/tournament/tournament-ranking/LCK%20Spring%202024/",
        "used_for": "Cross-check regular-season game W/L",
        "note": "",
    },
    {
        "id": "S13",
        "name": "gol.gg LCK 2024 Summer ranking",
        "url": "https://gol.gg/tournament/tournament-ranking/LCK%20Summer%202024/",
        "used_for": "Cross-check regular-season game W/L",
        "note": "",
    },
    {
        "id": "S14",
        "name": "gol.gg LCK 2025 Rounds 1-2 ranking",
        "url": "https://gol.gg/tournament/tournament-ranking/LCK%202025%20Rounds%201-2/",
        "used_for": "Cross-check regular-season game W/L",
        "note": "",
    },
    {
        "id": "S15",
        "name": "gol.gg LCK 2025 Rounds 3-5 ranking",
        "url": "https://gol.gg/tournament/tournament-ranking/LCK%202025%20Rounds%203-5/",
        "used_for": "Stage-only game W/L cross-check",
        "note": "gol.gg lists R3-5 stage-only game records; Leaguepedia standings carry over R1-2.",
    },
    {
        "id": "S16",
        "name": "gol.gg LCK 2026 Rounds 1-2 ranking",
        "url": "https://gol.gg/tournament/tournament-ranking/LCK%202026%20Rounds%201-2/",
        "used_for": "Cross-check current 2026 game W/L",
        "note": "gol.gg ranking is game-record sorted; Leaguepedia official table is series-record sorted.",
    },
    {
        "id": "S17",
        "name": "LPL 2025 Split 2 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LPL/2025_Season/Split_2",
        "used_for": "Peyz 2025 LPL Split 2 All-Pro",
        "note": "",
    },
    {
        "id": "S18",
        "name": "LPL 2025 Split 3 - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LPL/2025_Season/Split_3",
        "used_for": "Negative check: Peyz is not listed on LPL 2025 Split 3 All-Pro teams",
        "note": "",
    },
    {
        "id": "S19",
        "name": "LPL 2026 Split 1 Playoffs - Leaguepedia",
        "url": "https://lol.fandom.com/wiki/LPL/2026_Season/Split_1_Playoffs",
        "used_for": "Viper/BLG playoff placement context only",
        "note": "Not a regular-league row for first-review league scoring.",
    },
]


PLAYERS = [
    ("Zeus", "Choi Woo-je", "TOP", "selected", "top", "2024 T1; 2025-2026 HLE"),
    ("Kiin", "Kim Gi-in", "TOP", "contender", "top", "2024-2026 Gen.G"),
    ("Canyon", "Kim Geon-bu", "JNG", "selected", "jungle", "2024-2026 Gen.G"),
    ("Oner", "Mun Hyeon-jun", "JNG", "contender", "jungle", "2024-2026 T1"),
    ("Faker", "Lee Sang-hyeok", "MID", "selected", "mid/context", "2024-2026 T1"),
    ("Zeka", "Kim Geon-woo", "MID", "selected", "mid/context", "2024-2026 HLE"),
    ("Chovy", "Jeong Ji-hoon", "MID", "context", "mid/context", "2024-2026 Gen.G"),
    ("Gumayusi", "Lee Min-hyeong", "BOT", "selected", "bot", "2024-2025 T1; 2025 KeSPA/2026 HLE"),
    ("Peyz", "Kim Su-hwan", "BOT", "contender", "bot", "2024 Gen.G; 2025 JDG; 2025 KeSPA/2026 T1"),
    ("Viper", "Park Do-hyeon", "BOT", "contender", "bot", "2024-2025 HLE; 2026 BLG"),
    ("Ruler", "Park Jae-hyuk", "BOT", "context", "bot/context", "2025-2026 Gen.G"),
    ("Keria", "Ryu Min-seok", "SUP", "selected", "support/context", "2024-2026 T1"),
    ("Doran", "Choi Hyeon-joon", "TOP", "context", "top/context", "2024 HLE; 2025-2026 T1"),
    ("Peanut", "Han Wang-ho", "JNG", "context", "jungle/context", "2024-2025 HLE"),
    ("Kanavi", "Seo Jin-hyeok", "JNG", "context", "jungle/context", "2026 HLE"),
    ("Delight", "Yoo Hwan-joong", "SUP", "context", "support/context", "2024-2026 HLE"),
]


ROSTERS = {
    "GEN_2024": ["Kiin", "Canyon", "Chovy", "Peyz"],
    "T1_2024": ["Zeus", "Oner", "Faker", "Gumayusi", "Keria"],
    "HLE_2024": ["Doran", "Peanut", "Zeka", "Viper", "Delight"],
    "GEN_2025": ["Kiin", "Canyon", "Chovy", "Ruler"],
    "T1_2025": ["Doran", "Oner", "Faker", "Gumayusi", "Keria"],
    "HLE_2025": ["Zeus", "Peanut", "Zeka", "Viper", "Delight"],
    "GEN_2026": ["Kiin", "Canyon", "Chovy", "Ruler"],
    "T1_2026": ["Doran", "Oner", "Faker", "Peyz", "Keria"],
    "HLE_2026": ["Zeus", "Kanavi", "Zeka", "Gumayusi", "Delight"],
    "JDG_2025": ["Peyz"],
    "BLG_2026": ["Viper"],
}


TEAM_PLAYERS = {
    ("2024 LCK Spring Regular", "Gen.G"): ROSTERS["GEN_2024"],
    ("2024 LCK Spring Regular", "T1"): ROSTERS["T1_2024"],
    ("2024 LCK Spring Regular", "HLE"): ROSTERS["HLE_2024"],
    ("2024 LCK Summer Regular", "Gen.G"): ROSTERS["GEN_2024"],
    ("2024 LCK Summer Regular", "T1"): ROSTERS["T1_2024"],
    ("2024 LCK Summer Regular", "HLE"): ROSTERS["HLE_2024"],
    ("2025 LCK Rounds 1-2 Regular", "Gen.G"): ROSTERS["GEN_2025"],
    ("2025 LCK Rounds 1-2 Regular", "T1"): ROSTERS["T1_2025"],
    ("2025 LCK Rounds 1-2 Regular", "HLE"): ROSTERS["HLE_2025"],
    ("2025 LCK Rounds 3-5 Regular", "Gen.G"): ROSTERS["GEN_2025"],
    ("2025 LCK Rounds 3-5 Regular", "T1"): ROSTERS["T1_2025"],
    ("2025 LCK Rounds 3-5 Regular", "HLE"): ROSTERS["HLE_2025"],
    ("2026 LCK Rounds 1-2 Regular Snapshot", "Gen.G"): ROSTERS["GEN_2026"],
    ("2026 LCK Rounds 1-2 Regular Snapshot", "T1"): ROSTERS["T1_2026"],
    ("2026 LCK Rounds 1-2 Regular Snapshot", "HLE"): ROSTERS["HLE_2026"],
    ("2025 LPL Split 1 Regular", "JDG"): ROSTERS["JDG_2025"],
    ("2026 LPL Split 1", "BLG"): ROSTERS["BLG_2026"],
}


LEAGUE_STANDINGS = [
    ("2024 LCK Spring Regular", "LCK", "2024 Spring", "Gen.G", 1, 17, 1, 34, 5, 29, "S2", "regular season; game W/L cross-checked with S12"),
    ("2024 LCK Spring Regular", "LCK", "2024 Spring", "T1", 2, 15, 3, 32, 8, 24, "S2", "regular season; game W/L cross-checked with S12"),
    ("2024 LCK Spring Regular", "LCK", "2024 Spring", "HLE", 3, 15, 3, 30, 11, 19, "S2", "regular season; game W/L cross-checked with S12"),
    ("2024 LCK Summer Regular", "LCK", "2024 Summer", "Gen.G", 1, 17, 1, 35, 3, 32, "S3", "regular season; game W/L cross-checked with S13"),
    ("2024 LCK Summer Regular", "LCK", "2024 Summer", "HLE", 2, 14, 4, 30, 11, 19, "S3", "regular season; game W/L cross-checked with S13"),
    ("2024 LCK Summer Regular", "LCK", "2024 Summer", "T1", 4, 11, 7, 25, 19, 6, "S3", "regular season; game W/L cross-checked with S13"),
    ("2025 LCK Rounds 1-2 Regular", "LCK", "2025 R1-2", "Gen.G", 1, 18, 0, 36, 5, 31, "S4", "regular season; game W/L cross-checked with S14"),
    ("2025 LCK Rounds 1-2 Regular", "LCK", "2025 R1-2", "HLE", 2, 14, 4, 31, 11, 20, "S4", "regular season; game W/L cross-checked with S14"),
    ("2025 LCK Rounds 1-2 Regular", "LCK", "2025 R1-2", "T1", 3, 11, 7, 25, 17, 8, "S4", "regular season; game W/L cross-checked with S14"),
    ("2025 LCK Rounds 3-5 Regular", "LCK", "2025 R3-5", "Gen.G", 1, 29, 1, 59, 11, 48, "S5", "records carried over from R1-2"),
    ("2025 LCK Rounds 3-5 Regular", "LCK", "2025 R3-5", "HLE", 2, 20, 10, 46, 24, 22, "S5", "records carried over from R1-2"),
    ("2025 LCK Rounds 3-5 Regular", "LCK", "2025 R3-5", "T1", 3, 20, 10, 45, 26, 19, "S5", "records carried over from R1-2"),
    ("2026 LCK Rounds 1-2 Regular Snapshot", "LCK", "2026 R1-2 through Week 7", "HLE", 1, 12, 2, 25, 8, 17, "S6", "snapshot before full R1-2 completion"),
    ("2026 LCK Rounds 1-2 Regular Snapshot", "LCK", "2026 R1-2 through Week 7", "T1", 3, 10, 4, 22, 9, 13, "S6", "snapshot before full R1-2 completion"),
    ("2026 LCK Rounds 1-2 Regular Snapshot", "LCK", "2026 R1-2 through Week 7", "Gen.G", 4, 10, 4, 22, 10, 12, "S6", "snapshot before full R1-2 completion"),
    ("2025 LPL Split 1 Regular", "LPL", "2025 Split 1", "JDG", 1, 3, 0, 9, 4, 5, "S9", "regular group-stage row only; Split 2/3 regular rows still missing"),
    ("2026 LPL Split 1", "LPL", "2026 Split 1", "BLG", 2, 6, 4, 15, 10, 5, "S10", "regular group-stage row; BLG later won playoffs per S19"),
]


GOLGG_STAGE_ONLY = [
    ("2024 LCK Spring Regular", "Gen.G", 34, 5, 29, "S12", "matches Leaguepedia regular games"),
    ("2024 LCK Spring Regular", "T1", 32, 8, 24, "S12", "matches Leaguepedia regular games"),
    ("2024 LCK Spring Regular", "HLE", 30, 11, 19, "S12", "matches Leaguepedia regular games"),
    ("2024 LCK Summer Regular", "Gen.G", 35, 3, 32, "S13", "matches Leaguepedia regular games"),
    ("2024 LCK Summer Regular", "HLE", 30, 11, 19, "S13", "matches Leaguepedia regular games"),
    ("2024 LCK Summer Regular", "T1", 25, 19, 6, "S13", "matches Leaguepedia regular games"),
    ("2025 LCK Rounds 1-2 Regular", "Gen.G", 36, 5, 31, "S14", "matches Leaguepedia regular games"),
    ("2025 LCK Rounds 1-2 Regular", "HLE", 31, 11, 20, "S14", "matches Leaguepedia regular games"),
    ("2025 LCK Rounds 1-2 Regular", "T1", 25, 17, 8, "S14", "matches Leaguepedia regular games"),
    ("2025 LCK Rounds 3-5 stage only", "Gen.G", 23, 6, 17, "S15", "gol.gg stage-only; Leaguepedia standings carry over R1-2"),
    ("2025 LCK Rounds 3-5 stage only", "HLE", 15, 13, 2, "S15", "gol.gg stage-only; Leaguepedia standings carry over R1-2"),
    ("2025 LCK Rounds 3-5 stage only", "T1", 20, 9, 11, "S15", "gol.gg stage-only; Leaguepedia standings carry over R1-2"),
    ("2026 LCK Rounds 1-2 Snapshot", "HLE", 25, 8, 17, "S16", "games match Leaguepedia; ranking display may differ by series vs game order"),
    ("2026 LCK Rounds 1-2 Snapshot", "T1", 22, 9, 13, "S16", "games match Leaguepedia; ranking display may differ by series vs game order"),
    ("2026 LCK Rounds 1-2 Snapshot", "Gen.G", 22, 10, 12, "S16", "games match Leaguepedia; ranking display may differ by series vs game order"),
]


TOURNAMENT_RESULTS = [
    ("2024 Worlds", "Worlds", "T1", "1st", 1.0, ROSTERS["T1_2024"], "S11", ""),
    ("2024 Worlds", "Worlds", "Gen.G", "3rd-4th", 3.5, ROSTERS["GEN_2024"], "S11", ""),
    ("2024 Worlds", "Worlds", "HLE", "5th-8th", 5.5, ROSTERS["HLE_2024"], "S11", ""),
    ("2025 Worlds", "Worlds", "T1", "1st", 1.0, ROSTERS["T1_2025"], "S11", ""),
    ("2025 Worlds", "Worlds", "Gen.G", "3rd-4th", 3.5, ROSTERS["GEN_2025"], "S11", ""),
    ("2025 Worlds", "Worlds", "HLE", "5th-8th", 5.5, ROSTERS["HLE_2025"], "S11", ""),
    ("2024 KeSPA Cup", "KeSPA", "Gen.G", "3rd-4th", 3.5, ["Kiin", "Canyon", "Chovy"], "S8", "roster mixed; only tracked candidates/context included"),
    ("2024 KeSPA Cup", "KeSPA", "HLE", "3rd-4th", 3.5, ["Zeka", "Viper", "Delight"], "S8", "roster mixed; only tracked candidates/context included"),
    ("2024 KeSPA Cup", "KeSPA", "T1", "11th-12th", 11.5, ["Oner", "Gumayusi"], "S8", "T1 fielded mixed/academy roster; tracked players only"),
    ("2025 KeSPA Cup", "KeSPA", "T1", "1st", 1.0, ["Doran", "Oner", "Faker", "Peyz", "Keria"], "S7", ""),
    ("2025 KeSPA Cup", "KeSPA", "HLE", "2nd", 2.0, ["Zeus", "Kanavi", "Zeka", "Gumayusi", "Delight"], "S7", ""),
    ("2025 KeSPA Cup", "KeSPA", "Gen.G", "13th-14th / group 5th", 13.5, ["Kiin", "Canyon"], "S7", "Inven summary uses group-place style for Gen.G; Leaguepedia overall is 13th-14th."),
]


ALL_PRO = [
    ("2024 LCK Spring", "Kiin", "TOP", "Gen.G", 1, "S2"),
    ("2024 LCK Spring", "Canyon", "JNG", "Gen.G", 1, "S2"),
    ("2024 LCK Spring", "Chovy", "MID", "Gen.G", 1, "S2"),
    ("2024 LCK Spring", "Peyz", "BOT", "Gen.G", 1, "S2"),
    ("2024 LCK Spring", "Keria", "SUP", "T1", 1, "S2"),
    ("2024 LCK Spring", "Zeus", "TOP", "T1", 2, "S2"),
    ("2024 LCK Spring", "Oner", "JNG", "T1", 2, "S2"),
    ("2024 LCK Spring", "Faker", "MID", "T1", 2, "S2"),
    ("2024 LCK Spring", "Viper", "BOT", "HLE", 2, "S2"),
    ("2024 LCK Spring", "Doran", "TOP", "HLE", 3, "S2"),
    ("2024 LCK Spring", "Peanut", "JNG", "HLE", 3, "S2"),
    ("2024 LCK Spring", "Zeka", "MID", "HLE", 3, "S2"),
    ("2024 LCK Spring", "Gumayusi", "BOT", "T1", 3, "S2"),
    ("2024 LCK Spring", "Delight", "SUP", "HLE", 3, "S2"),
    ("2024 LCK Summer", "Kiin", "TOP", "Gen.G", 1, "S3"),
    ("2024 LCK Summer", "Canyon", "JNG", "Gen.G", 1, "S3"),
    ("2024 LCK Summer", "Chovy", "MID", "Gen.G", 1, "S3"),
    ("2024 LCK Summer", "Peyz", "BOT", "Gen.G", 1, "S3"),
    ("2024 LCK Summer", "Doran", "TOP", "HLE", 2, "S3"),
    ("2024 LCK Summer", "Peanut", "JNG", "HLE", 2, "S3"),
    ("2024 LCK Summer", "Zeka", "MID", "HLE", 2, "S3"),
    ("2024 LCK Summer", "Viper", "BOT", "HLE", 2, "S3"),
    ("2024 LCK Summer", "Delight", "SUP", "HLE", 2, "S3"),
    ("2024 LCK Summer", "Oner", "JNG", "T1", 3, "S3"),
    ("2024 LCK Summer", "Keria", "SUP", "T1", 3, "S3"),
    ("2025 LCK Rounds 3-5", "Kiin", "TOP", "Gen.G", 1, "S5"),
    ("2025 LCK Rounds 3-5", "Canyon", "JNG", "Gen.G", 1, "S5"),
    ("2025 LCK Rounds 3-5", "Chovy", "MID", "Gen.G", 1, "S5"),
    ("2025 LCK Rounds 3-5", "Ruler", "BOT", "Gen.G", 1, "S5"),
    ("2025 LCK Rounds 3-5", "Doran", "TOP", "T1", 2, "S5"),
    ("2025 LCK Rounds 3-5", "Oner", "JNG", "T1", 2, "S5"),
    ("2025 LCK Rounds 3-5", "Viper", "BOT", "HLE", 2, "S5"),
    ("2025 LCK Rounds 3-5", "Keria", "SUP", "T1", 2, "S5"),
    ("2025 LCK Rounds 3-5", "Zeus", "TOP", "HLE", 3, "S5"),
    ("2025 LCK Rounds 3-5", "Peanut", "JNG", "HLE", 3, "S5"),
    ("2025 LCK Rounds 3-5", "Faker", "MID", "T1", 3, "S5"),
    ("2025 LCK Rounds 3-5", "Gumayusi", "BOT", "T1", 3, "S5"),
    ("2025 LCK Rounds 3-5", "Delight", "SUP", "HLE", 3, "S5"),
    ("2025 LPL Split 1", "Peyz", "BOT", "JDG", 3, "S9"),
    ("2025 LPL Split 2", "Peyz", "BOT", "JDG", 2, "S17"),
    ("2026 LPL Split 1", "Viper", "BOT", "BLG", 2, "S10"),
]


VALIDATION_REPORT = [
    (
        "LCK standings",
        "corrected",
        "Filled missing game W/L for 2024 Spring, 2024 Summer, 2025 R1-2, and HLE 2025 R3-5.",
        "Leaguepedia rows are now cross-checkable against gol.gg rows in GOLGG_StageOnly_Check.",
    ),
    (
        "2025 LCK Rounds 3-5",
        "definition caveat",
        "Leaguepedia standings carry over R1-2 records, while gol.gg tournament ranking shows R3-5 stage-only game records.",
        "Keep cumulative standings in League_Standings and stage-only records in GOLGG_StageOnly_Check.",
    ),
    (
        "2026 LCK Rounds 1-2 snapshot",
        "definition caveat",
        "Game W/L match gol.gg, but ranking display can differ because Leaguepedia uses official series standings and gol.gg sorts game records.",
        "Use Leaguepedia rank for official league placement; use gol.gg only for game W/L checks.",
    ),
    (
        "All-Pro",
        "corrected",
        "Added missing tracked context rows for Chovy, Doran, Peanut, and Delight in LCK All-Pro sheets.",
        "AllPro is still tracked-player scoped, not a league-wide awards archive.",
    ),
    (
        "Peyz LPL All-Pro",
        "corrected",
        "Removed erroneous 2025 LPL Split 3 1st All-Pro row; Leaguepedia lists JackeyLove/GALA/Elk as Split 3 All-Pro ADCs, not Peyz.",
        "Keep Peyz Split 1 3rd and Split 2 2nd only unless another verified source is added.",
    ),
    (
        "LPL regular rows",
        "corrected/partial",
        "Changed Viper/BLG 2026 Split 1 league row from playoff champion rank 1 to regular Group Ascend rank 2; added JDG/Peyz 2025 Split 1 regular row.",
        "Do not run final Gumayusi-Peyz-Viper weighting until missing LPL Split 2/3 regular and POG/POM rows are completed.",
    ),
    (
        "LPL POG/POM",
        "incomplete",
        "LPL POG/POM/MVP rows for Peyz and Viper are not collected in POG_POM.",
        "Candidate_Summary currently undercounts Peyz/Viper POG/POM.",
    ),
    (
        "POG/POM scope",
        "partial",
        "POG_POM is candidate-focused and not a full league-wide POG archive; some context-player values are intentionally absent.",
        "Use source pages directly or expand POG_POM before doing full-league normalization or cross-position rankings.",
    ),
    (
        "Detailed stats",
        "incomplete",
        "KDA, KP, DPM, GD@14, vision and other detailed player stats are not collected.",
        "Workbook is usable for first-review public indices only, not final detailed-stat review.",
    ),
]


POG_POM = [
    ("2024 LCK Spring", "POG points", "Zeus", "T1", 600, 600, "S2"),
    ("2024 LCK Spring", "POG points", "Kiin", "Gen.G", 400, 400, "S2"),
    ("2024 LCK Spring", "POG points", "Canyon", "Gen.G", 1000, 1000, "S2"),
    ("2024 LCK Spring", "POG points", "Oner", "T1", 600, 600, "S2"),
    ("2024 LCK Spring", "POG points", "Faker", "T1", 1100, 1100, "S2"),
    ("2024 LCK Spring", "POG points", "Zeka", "HLE", 1000, 1000, "S2"),
    ("2024 LCK Spring", "POG points", "Gumayusi", "T1", 300, 300, "S2"),
    ("2024 LCK Spring", "POG points", "Peyz", "Gen.G", 200, 200, "S2"),
    ("2024 LCK Spring", "POG points", "Viper", "HLE", 600, 600, "S2"),
    ("2024 LCK Spring", "POG points", "Keria", "T1", 600, 600, "S2"),
    ("2024 LCK Summer", "POG points", "Zeus", "T1", 500, 500, "S3"),
    ("2024 LCK Summer", "POG points", "Kiin", "Gen.G", 500, 500, "S3"),
    ("2024 LCK Summer", "POG points", "Canyon", "Gen.G", 1000, 1000, "S3"),
    ("2024 LCK Summer", "POG points", "Oner", "T1", 600, 600, "S3"),
    ("2024 LCK Summer", "POG points", "Faker", "T1", 600, 600, "S3"),
    ("2024 LCK Summer", "POG points", "Zeka", "HLE", 400, 400, "S3"),
    ("2024 LCK Summer", "POG points", "Gumayusi", "T1", 400, 400, "S3"),
    ("2024 LCK Summer", "POG points", "Peyz", "Gen.G", 1100, 1100, "S3"),
    ("2024 LCK Summer", "POG points", "Viper", "HLE", 400, 400, "S3"),
    ("2024 LCK Summer", "POG points", "Keria", "T1", 400, 400, "S3"),
    ("2025 LCK Rounds 1-2", "POM times", "Zeus", "HLE", 7, 700, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Kiin", "Gen.G", 2, 200, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Canyon", "Gen.G", 1, 100, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Oner", "T1", 5, 500, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Faker", "T1", 2, 200, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Zeka", "HLE", 3, 300, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Gumayusi", "T1", 1, 100, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Viper", "HLE", 3, 300, "S4"),
    ("2025 LCK Rounds 1-2", "POM times", "Keria", "T1", 2, 200, "S4"),
    ("2025 LCK Rounds 3-5", "POM times", "Zeus", "HLE", 1, 100, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Kiin", "Gen.G", 3, 300, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Canyon", "Gen.G", 1, 100, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Gumayusi", "T1", 1, 100, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Viper", "HLE", 4, 400, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Keria", "T1", 6, 600, "S5"),
    ("2025 LCK Rounds 3-5", "POM times", "Ruler", "Gen.G", 4, 400, "S5"),
    ("2025 KeSPA Cup", "POG points", "Zeus", "HLE", 3, 3, "S7"),
    ("2025 KeSPA Cup", "POG points", "Oner", "T1", 1, 1, "S7"),
    ("2025 KeSPA Cup", "POG points", "Gumayusi", "HLE", 2, 2, "S7"),
    ("2025 KeSPA Cup", "POG points", "Peyz", "T1", 4, 4, "S7"),
    ("2025 KeSPA Cup", "POG points", "Faker", "T1", 0, 0, "S7"),
    ("2025 KeSPA Cup", "POG points", "Keria", "T1", 5, 5, "S7"),
    ("2025 KeSPA Cup", "POG points", "Zeka", "HLE", 2, 2, "S7"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Zeus", "HLE", 200, 200, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Kiin", "Gen.G", 200, 200, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Canyon", "Gen.G", 100, 100, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Oner", "T1", 300, 300, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Faker", "T1", 200, 200, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Zeka", "HLE", 500, 500, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Gumayusi", "HLE", 100, 100, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Peyz", "T1", 100, 100, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Keria", "T1", 300, 300, "S6"),
    ("2026 LCK Rounds 1-2 Snapshot", "POM points", "Kanavi", "HLE", 200, 200, "S6"),
]


def placement_score(rank: float | int | None) -> int | None:
    if rank is None:
        return None
    if rank <= 1:
        return 100
    if rank <= 2:
        return 85
    if rank <= 3:
        return 75
    if rank <= 4:
        return 65
    if rank <= 6:
        return 55
    if rank <= 8:
        return 45
    if rank <= 12:
        return 25
    return 10


def allpro_score(team: int) -> int:
    return {1: 3, 2: 2, 3: 1}.get(team, 0)


def append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def autosize(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(55, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    readme_rows = [
        ("Field", "Value"),
        ("Purpose", "Asian Games LoL selection-criteria dataset for follow-up calculations."),
        ("Created", "2026-05-19 KST"),
        ("Validated", "2026-05-19 KST against Leaguepedia/lol.fandom.com and gol.gg public pages."),
        ("Scope", "Publicly available first-review criteria: regional league results, Worlds, KeSPA Cup, All-Pro, POG/POM."),
        ("Players", "Selected roster + major disputed contenders + context players."),
        ("Important caveat", "Official weights are not public; scoring maps and weights are editable and not official."),
        ("Validation caveat", "LPL rows and detailed personal stats remain partial; see Validation_Report before using bot-lane comparisons."),
        ("Detailed stat caveat", "KDA, KP, DPM, GD@14, vision etc. are listed as future-stat fields but not fully collected here."),
    ]
    append_rows(readme, ["Field", "Value"], readme_rows[1:])

    sources = wb.create_sheet("Sources")
    append_rows(
        sources,
        ["source_id", "name", "url", "used_for", "note"],
        [(s["id"], s["name"], s["url"], s["used_for"], s["note"]) for s in SOURCES],
    )

    players = wb.create_sheet("Players")
    append_rows(
        players,
        ["player", "real_name", "position", "selection_status", "comparison_group", "teams_in_scope"],
        PLAYERS,
    )

    league = wb.create_sheet("League_Standings")
    append_rows(
        league,
        [
            "event",
            "region",
            "period",
            "team",
            "rank",
            "series_w",
            "series_l",
            "games_w",
            "games_l",
            "game_diff",
            "default_place_score",
            "source_id",
            "note",
        ],
        [
            (
                event,
                region,
                period,
                team,
                rank,
                sw,
                sl,
                gw,
                gl,
                gd,
                placement_score(rank),
                source,
                note,
            )
            for event, region, period, team, rank, sw, sl, gw, gl, gd, source, note in LEAGUE_STANDINGS
        ],
    )

    league_player_rows = []
    for event, region, period, team, rank, sw, sl, gw, gl, gd, source, note in LEAGUE_STANDINGS:
        for player in TEAM_PLAYERS.get((event, team), []):
            league_player_rows.append(
                (
                    event,
                    region,
                    period,
                    player,
                    team,
                    rank,
                    sw,
                    sl,
                    gw,
                    gl,
                    gd,
                    placement_score(rank),
                    source,
                    note,
                )
            )
    league_players = wb.create_sheet("League_By_Player")
    append_rows(
        league_players,
        [
            "event",
            "region",
            "period",
            "player",
            "team",
            "rank",
            "series_w",
            "series_l",
            "games_w",
            "games_l",
            "game_diff",
            "default_place_score",
            "source_id",
            "note",
        ],
        league_player_rows,
    )

    golgg_check = wb.create_sheet("GOLGG_StageOnly_Check")
    append_rows(
        golgg_check,
        ["event", "team", "golgg_games_w", "golgg_games_l", "golgg_game_diff", "source_id", "note"],
        GOLGG_STAGE_ONLY,
    )

    tr_rows = []
    for event, category, team, label, rank, roster, source, note in TOURNAMENT_RESULTS:
        for player in roster:
            tr_rows.append((event, category, player, team, label, rank, placement_score(rank), source, note))
    tournaments = wb.create_sheet("Worlds_KeSPA")
    append_rows(
        tournaments,
        ["event", "category", "player", "team", "placement", "rank_numeric", "default_place_score", "source_id", "note"],
        tr_rows,
    )

    allpro = wb.create_sheet("AllPro")
    append_rows(
        allpro,
        ["event", "player", "position", "team", "all_pro_team", "default_allpro_points", "source_id"],
        [(event, p, pos, team, tier, allpro_score(tier), source) for event, p, pos, team, tier, source in ALL_PRO],
    )

    pog = wb.create_sheet("POG_POM")
    append_rows(
        pog,
        ["event", "unit", "player", "team", "raw_value", "points_equiv", "source_id"],
        POG_POM,
    )

    validation = wb.create_sheet("Validation_Report")
    append_rows(
        validation,
        ["area", "status", "finding", "action"],
        VALIDATION_REPORT,
    )

    detailed = wb.create_sheet("Detailed_Stats_TODO")
    append_rows(
        detailed,
        ["position", "stat_name_reported", "collected", "note"],
        [
            ("ALL", "KDA", "no", "Reported as reviewed after first-review cut; not fully collected in this workbook."),
            ("ALL", "solo_kill_average", "no", ""),
            ("ALL", "KP", "no", "Kill participation."),
            ("ALL", "champion_count", "no", ""),
            ("ALL", "DMG%", "no", ""),
            ("ALL", "DPM", "no", ""),
            ("ALL", "DPG", "no", ""),
            ("ALL", "GDPM", "no", ""),
            ("ALL", "14CSD", "no", "CS differential at 14."),
            ("ALL", "14DPG", "no", "Damage/gold differential at 14 as reported; exact public field names differ by site."),
            ("ALL", "TeamGold", "no", ""),
            ("ALL", "TeamDMG", "no", ""),
            ("ALL", "vision_score", "no", ""),
            ("JNG", "first_blood_contribution", "no", ""),
            ("JNG", "enemy_jungle_time", "no", ""),
            ("JNG", "counter_jungle_cs", "no", ""),
            ("BOT", "duo_proximity", "no", ""),
            ("SUP", "duo_proximity", "no", ""),
            ("SUP", "jungle_proximity", "no", ""),
        ],
    )

    weights = wb.create_sheet("Weights")
    append_rows(
        weights,
        ["component", "weight_editable", "note"],
        [
            ("league", 0.17, "User-pasted max-min example; editable."),
            ("worlds", 0.26, "User-pasted max-min example; editable."),
            ("kespa", 0.05, "User-pasted max-min example; editable."),
            ("allpro", 0.05, "User-pasted max-min example; editable."),
            ("pog_pom", 0.47, "User-pasted max-min example; editable."),
        ],
    )
    weights["E1"] = "Default placement score map"
    weights["E2"] = "rank<=1"
    weights["F2"] = 100
    weights["E3"] = "rank<=2"
    weights["F3"] = 85
    weights["E4"] = "rank<=3"
    weights["F4"] = 75
    weights["E5"] = "rank<=4"
    weights["F5"] = 65
    weights["E6"] = "rank<=6"
    weights["F6"] = 55
    weights["E7"] = "rank<=8"
    weights["F7"] = 45
    weights["E8"] = "rank<=12"
    weights["F8"] = 25
    weights["E9"] = "else"
    weights["F9"] = 10

    summary_rows = []
    tracked = [p[0] for p in PLAYERS if p[3] in {"selected", "contender"}]
    league_scores = defaultdict(list)
    worlds_scores = defaultdict(list)
    kespa_scores = defaultdict(list)
    allpro_scores = defaultdict(int)
    pog_scores = defaultdict(int)

    for row in league_player_rows:
        player = row[3]
        score = row[11]
        if score is not None:
            league_scores[player].append(score)

    for row in tr_rows:
        _, category, player, *_rest = row
        score = row[6]
        if category == "Worlds":
            worlds_scores[player].append(score)
        elif category == "KeSPA":
            kespa_scores[player].append(score)

    for event, p, pos, team, tier, source in ALL_PRO:
        allpro_scores[p] += allpro_score(tier)

    for event, unit, player, team, raw, points, source in POG_POM:
        pog_scores[player] += points or 0

    def avg(values):
        return round(sum(values) / len(values), 2) if values else 0

    max_allpro = max((allpro_scores[p] for p in tracked), default=1)
    max_pog = max((pog_scores[p] for p in tracked), default=1)

    for player in tracked:
        l = avg(league_scores[player])
        w = avg(worlds_scores[player])
        k = avg(kespa_scores[player])
        a_raw = allpro_scores[player]
        p_raw = pog_scores[player]
        a_norm = round(a_raw / max_allpro * 100, 2) if max_allpro else 0
        p_norm = round(p_raw / max_pog * 100, 2) if max_pog else 0
        summary_rows.append(
            [
                player,
                l,
                w,
                k,
                a_raw,
                a_norm,
                p_raw,
                p_norm,
                None,
                "Example only: default score maps; normalization is selected/contender scoped; not official.",
            ]
        )

    summary = wb.create_sheet("Candidate_Summary")
    append_rows(
        summary,
        [
            "player",
            "league_score_avg",
            "worlds_score_avg",
            "kespa_score_avg",
            "allpro_points_raw",
            "allpro_norm_0_100",
            "pog_pom_points_raw",
            "pog_pom_norm_0_100",
            "weighted_total_formula",
            "note",
        ],
        summary_rows,
    )

    # Add formulas after writing rows. Weight cells: B2:B6.
    for r in range(2, summary.max_row + 1):
        summary.cell(r, 9).value = f"=B{r}*Weights!$B$2+C{r}*Weights!$B$3+D{r}*Weights!$B$4+F{r}*Weights!$B$5+H{r}*Weights!$B$6"

    for ws in wb.worksheets:
        autosize(ws)

    wb.save(OUT)

    # Lightweight verification that the generated workbook opens.
    loaded = load_workbook(OUT, data_only=False, read_only=True)
    print(f"wrote {OUT.resolve()}")
    print(f"sheets: {', '.join(loaded.sheetnames)}")
    print(f"rows Candidate_Summary={loaded['Candidate_Summary'].max_row}")
    loaded.close()


if __name__ == "__main__":
    main()
